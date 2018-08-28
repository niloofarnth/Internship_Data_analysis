# -*- coding: utf-8 -*-
"""
Created on Wed Aug  1 14:55:00 2018

@author: nxtehr
"""

import pandas as pd
import numpy as np
import os
import re
import math
from scipy import stats
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import Savitzky_Golay_Filtering
import basic_functions


''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
def clean_data(file_name):
    path = 'J:/nxtehr/Lab Results/Tensile test/Q&P tensile test files/'

    xls = pd.ExcelFile(path + file_name)
    wb = load_workbook(path + file_name)
    sheet_names = wb.sheetnames
    info = pd.read_excel(xls, 'Test Report')
    
    if info ['Default Test Run Report'].iloc[36] == 'Diameter':
        peak_load = info['Unnamed: 1'].iloc[43]
        yield_stress = info['Unnamed: 1'].iloc[46]
        modulus = info['Unnamed: 1'].iloc[45]
    else:   
        peak_load = info['Unnamed: 1'].iloc[42]
        yield_stress = info['Unnamed: 1'].iloc[45]
        modulus = info['Unnamed: 1'].iloc[44]
    
    if 'Raw Data'in sheet_names:
        sheet_name = 'Raw Data'
    elif 'raw data' in sheet_names:
        sheet_name = 'raw data'
    elif 'Sheet1' in sheet_names:
        sheet_name = 'Sheet1'   
    
    raw_data = pd.read_excel(xls, sheet_name)
    raw_data = raw_data[[raw_data.columns[3],raw_data.columns[4]]].iloc[6:].reset_index(drop = True)
    raw_data.columns = ['Strain (in/in)','Stress (ksi)']
    while isinstance(raw_data['Stress (ksi)'].iloc[0], str) or str(raw_data['Stress (ksi)'].iloc[0]) =='nan' :
        raw_data = raw_data.drop([0]).reset_index(drop=True)
    raw_data['Stress (MPa)'] =  raw_data['Stress (ksi)']*6.89475908677537   
    
    return (raw_data, yield_stress, modulus)
a = clean_data('5181-1.xlsx')
''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
#calculate the desired columns based on the standard E646-16 for total true strain
def add_columns_total_strain(file_name): 
    (raw_data, yield_stress, modulus) = clean_data(file_name)
    min_stress_index = raw_data.index[raw_data['Stress (ksi)']> yield_stress].tolist()[0]
    max_stress_index = raw_data.index[raw_data['Stress (ksi)'] == raw_data['Stress (ksi)'].max()].tolist()[0]
    UTS_strain = raw_data['Strain (in/in)'].iloc[max_stress_index]
    max_strain = pd.to_numeric(raw_data['Strain (in/in)'],errors = 'coerce').max()
    data_section = raw_data.iloc[min_stress_index:max_stress_index].reset_index(drop = True)
    data_section['true_stress(ksi)'] = data_section['Stress (ksi)']*(1+ pd.to_numeric(data_section['Strain (in/in)'],errors = 'coerce'))
    data_section['Y = log(true stress)'] = np.log10(np.float64(data_section['true_stress(ksi)']))
    data_section['Y^2'] = data_section['Y = log(true stress)']**2
    data_section['true_strain'] = np.log(1+ np.float64(data_section['Strain (in/in)']))
    data_section['X = log(true strain)']  = np.log10(np.float64(data_section['true_strain']))
    data_section['X^2'] = data_section['X = log(true strain)']**2
    data_section['X*Y'] = data_section['X = log(true strain)']* data_section['Y = log(true stress)'] 
    return data_section


''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
#calculate the desired columns based on the standard E646-16 for plastic true strain
def add_columns_plastic_strain(file_name): 
    (raw_data, yield_stress, modulus) = clean_data(file_name)
    min_stress_index = raw_data.index[raw_data['Stress (ksi)']> yield_stress].tolist()[0]
    max_stress_index = raw_data.index[raw_data['Stress (ksi)'] == raw_data['Stress (ksi)'].max()].tolist()[0]
    UTS_strain = raw_data['Strain (in/in)'].iloc[max_stress_index]
    max_strain = pd.to_numeric(raw_data['Strain (in/in)'],errors = 'coerce').max()
    data_section = raw_data.iloc[min_stress_index:max_stress_index].reset_index(drop = True)
    data_section['true_stress(ksi)'] = data_section['Stress (ksi)']*(1+ pd.to_numeric(data_section['Strain (in/in)'],errors = 'coerce'))
    data_section['Y = log(true stress)'] = np.log10(np.float64(data_section['true_stress(ksi)']))
    data_section['Y^2'] = data_section['Y = log(true stress)']**2
    elastic_strain = data_section['true_stress(ksi)']/modulus
    data_section['true_strain'] = np.log(1+ np.float64(data_section['Strain (in/in)']))
    data_section['X_tot = log(true strain)']  = np.log10(np.float64(data_section['true_strain']))
    data_section['true_plastic_strain'] = np.log(1+ np.float64(data_section['Strain (in/in)']))-elastic_strain
    data_section['X = log(true plastic strain)']  = np.log10(np.float64(data_section['true_plastic_strain']))
    data_section['X^2'] = data_section['X = log(true plastic strain)']**2
    data_section['X*Y'] = data_section['X = log(true plastic strain)']* data_section['Y = log(true stress)'] 
    return data_section

''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''

def instantanous_n(sample):
    data = add_columns_plastic_strain(sample)
    new_X_list = []
    n_plastic_list = []
    n_tot_list = []
    for i in range(len(data)-15):
        #sigma_Y is the change in the true stress for a step size
        sigma_Y = data['Y = log(true stress)'].iloc[i+15]-data['Y = log(true stress)'].iloc[i]
        #sigma_X is the change in the true plastic strain for a step size
        sigma_X_plastic = data['X = log(true plastic strain)'].iloc[i+15]-data['X = log(true plastic strain)'].iloc[i]
        #sigma_X is the change in the true total strain for a step size
        sigma_X_tot = data['X_tot = log(true strain)'].iloc[i+15]-data['X_tot = log(true strain)'].iloc[i]
        new_X_tot=  (float(data['true_strain'].iloc[i+15])+ float(data['true_strain'].iloc[i]))/2

        n_plastic = sigma_Y/sigma_X_plastic
        n_plastic_list.append(n_plastic)
        n_tot = sigma_Y/sigma_X_tot
        n_tot_list.append(n_tot)
        
        new_X_list.append(new_X_tot)
        i += 16
        
    yhat = Savitzky_Golay_Filtering.savitzky_golay(n_tot_list, 51, 3) # window size 51, polynomial order 3
    yaxis = pd.Series(yhat)
    xaxis = pd.Series(new_X_list)
    return (xaxis,yaxis,n_tot_list)

def instant_n_plot(samples,path):
    for file in samples:
        name = file.split('-')[0]
        xaxis,yaxis,n_tot_list= instantanous_n(file)
        data_section = add_columns_plastic_strain(file)
        data_section_tot = add_columns_total_strain(file)
        front = basic_functions.front_or_back('front',path)
        condition = front.loc[front['sample'] == int(name)]
        
        PT =int(condition['partitioning temperature °C'])
        QT = int(condition['quenching temperature °C'])
        Pt = int(condition['partitioning time (min)'])
        t = int(condition['Thickness (mm)'])
        PLM = float(condition['P(L.M.)'])
        f, ax = plt.subplots(figsize=(7,7), dpi=200)
        ax.set_title('Thickness = {}mm'.format(t),fontsize = 12)
        ax.plot(xaxis*100, yaxis,linewidth=1,alpha = 0.6, label = 'QT = {}°C\nPT = {}°C\nPt = {}min'.format(QT,PT,Pt))
        #ax.plot(xaxis, n_list,linewidth=1,alpha = 0.6,label = 'n_plastic')
        #ax.plot(xaxis, n_tot_list,linewidth=1,alpha = 0.6,label = 'n_total')
        #ax.plot(data_section['Strain (in/in)'], data_section['Stress (ksi)'],linewidth=1,alpha = 0.6,label = 'engineering')
        #ax.plot(data_section['true_plastic_strain'], data_section['true_stress(ksi)'],linewidth=1,alpha = 0.6,label = 'true plastic')
        
        #ax.plot(data_section['true_strain'], data_section['true_stress(ksi)'],linewidth=1,alpha = 0.6,label = 'true total')
        ax.set_xlabel('$\epsilon_t$')
        ax.set_ylabel('$n_{instantanous}$')
    
        plt.legend(loc = 'upper right', title = 'condition')

   
    
    
    
    
    